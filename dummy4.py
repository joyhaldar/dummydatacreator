from sqlalchemy import create_engine
import cx_Oracle
import pandas as pd
from pandas import ExcelWriter
import random
import string
import time
from datetime import datetime
import os
import xlwt
import xlrd
import openpyxl
from random import randint



writer = ExcelWriter('PythonExport.xlsx')

db = input("Enter DB name: ")

table = input('Enter table name: ')

host='indlin2332'
port=1521
sid= 'SINODS1' #'CNG1D' #'SINODS1'
user= db #'IBMCOGNOS_CS' #'ODS_DEV_DDS_R2'
password=db #'IBMCOGNOS_CS' #'ODS_DEV_DDS_R2'
sid = cx_Oracle.makedsn(host, port, sid=sid)

#engine = create_engine('oracle://ODS_DEV_REPL_ABP_R2:ODS_DEV_REPL_ABP_R2@SINODS1')

loe = ["'I'","'U'","'D'"]

def randomString(stringLength):
    """Generate a random string with the combination of lowercase and uppercase letters """
    letters = string.ascii_letters
    return ''.join(random.choice(letters) for i in range(stringLength))
	
def random_with_N_digits(n):
    range_start = 10**(n-1)
    range_end = (10**n)-1
    return randint(range_start, range_end)
	
def strTimeProp(start, end, format, prop):
    """Get a time at a proportion of a range of two formatted times.

    start and end should be strings specifying times formated in the
    given format (strftime-style), giving an interval [start, end].
    prop specifies how a proportion of the interval to be taken after
    start.  The returned time will be in the specified format.
    """

    stime = time.mktime(time.strptime(start, format))
    etime = time.mktime(time.strptime(end, format))

    ptime = stime + prop * (etime - stime)

    return time.strftime(format, time.localtime(ptime))


def randomDate(start, end, prop):
    return strTimeProp(start, end, '%m/%d/%Y', prop) # %I:%M %p', prop)
	
cstr = 'oracle://{user}:{password}@{sid}'.format(
    user=user,
    password=password,
    sid=sid
)

def datime(date_string):
	#date_string = '2009-11-29 03:17:25 PM'
	format = '%Y-%m-%d %I:%M:%S %p'
	my_date = datetime.strptime(date_string, format)

	# This prints '2009-11-29 03:17:25 AM'
	#print(my_date.strftime(format))
	print("'"+my_date.strftime(format)+"'")
	return("'"+my_date.strftime(format)+"'")

engine =  create_engine(
    cstr,
    convert_unicode=False,
    pool_recycle=10,
    pool_size=50,
    echo=True
)

#result = engine.execute('SELECT * FROM AGREEMENT_REPL')

#for row in result:
#    print(row)


con = engine.connect()
output = engine.execute("select COLUMN_NAME, DATA_TYPE, DATA_PRECISION, DATA_SCALE, DATA_LENGTH, NULLABLE from USER_TAB_COLUMNS where TABLE_NAME = '"+table+"' AND DATA_TYPE NOT IN ('ROWID','CLOB','BLOB')")
df = pd.DataFrame(output.fetchall())
df.columns = output.keys()

#print(df) #.head())
#newdf=df['spec']
#print(newdf)

#df.to_excel(writer,'Sheet1')
#writer.save()

#book = xlwt.Workbook()
#sheet = book.add_sheet("sheet1")

wb = openpyxl.Workbook() 
sheet = wb.active 

i=1
for index, row in df.iterrows():
	#print(row['column_name'], row['data_type'], row['data_precision'], row['data_scale'], row['data_length'], row['nullable'])
	if row['data_type'] in ('ROWID'):
		pass
	else:
		#sheet.write(0,i, row['column_name'])
		sheet.cell(row = 1, column = i).value = row['column_name']
		#sheet.write(1,i, row['data_type'])
		sheet.cell(row = 2, column = i).value = row['data_type']
		sheet.cell(row = 3, column = i).value = row['data_precision']
		sheet.cell(row = 4, column = i).value = row['data_scale']
		sheet.cell(row = 5, column = i).value = row['data_length']
		#sheet.write(2,i, row['nullable'])
		sheet.cell(row = 6, column = i).value = row['nullable']
		i = i+1

#book.save("joinlist.xls")
#for k in range(1, 10):
j = 1
#m = 4
for index, row in df.iterrows():
	#print(row['column_name'], row['data_type'], row['data_length'], row['nullable'])
	#if row['column_name']=='CUST_GRP_ID':
	#	for m in range(7,16):
	#		sheet.cell(row = m, column = j).value = '1'
	if row['data_type'] in ('VARCHAR', 'VARCHAR2', 'CHAR', 'NVARCHAR2'):
		a = int(row['data_length'])
		a1 = randomString(1)
		s1 = "'"+a1+"'"
		#print(s1)
		#sheet.write(3,j, "'I'")
		for m in range(7,16):
			if sheet.cell(row = 1, column = j).value == 'OP_CODE':
				sheet.cell(row = m, column = j).value = random.choice(loe)  #"'I'"
			else:
				sheet.cell(row = m, column = j).value = s1
		j = j+1
	elif row['data_type'] in ('NUMBER'):
		print(j)
		print(row['column_name'])
		print(sheet.cell(row = 3, column = j).value)
		print(sheet.cell(row = 4, column = j).value)
		pr = 1 # int(sheet.cell(row = 3, column = j).value)
		sc = 1 # int(sheet.cell(row = 4, column = j).value)
		#s2 = 'na' #random_with_N_digits(pr)#'2'#'1' #random.randint(1, 1000) #random.sample(range(1, 10), 1)
		#print(s2)
		#sheet.write(3,j, s2)
		if row['column_name']=='CUST_GRP_ID':
			for m in range(7,16):
				sheet.cell(row = m, column = j).value = '2'
			j=j+1
		else:
			for m in range(7,16):
				#pr = int(sheet.cell(row = 3, column = j).value)
				#sc = int(sheet.cell(row = 4, column = j).value)
				s2 = random_with_N_digits(1)#'2'#'1' #random.randint(1, 1000) #random.sample(range(1, 10), 1)
				#print(s2)
				sheet.cell(row = m, column = j).value = s2 #m-3
			j = j+1
	elif row['data_type'] in ('DATE'):
		s3 = "'01-Jan-19'" #randomDate("1/1/2018", "1/2/2018", random.random())
		#print(s3)
		#sheet.write(3,j, s3)
		for m in range(7,16):
			sheet.cell(row = m, column = j).value = s3
		j = j+1
	elif 'TIMESTAMP' in row['data_type']: # like '%TIMESTAMP%':	#in ('TIMESTAMP'):
		#s4 = datime('2017-1-1 03:17:25 PM')
		s4 = "'"+"01-FEB-17 04.39.54.000000000 PM"+"'"
		#print(s4)
		#sheet.write(3,j, s4)
		for m in range(7,16):
			sheet.cell(row = m, column = j).value = s4
		j = j+1
	elif row['data_type'] in ('ROWID'):
		pass
	elif row['data_type'] in ('CLOB'):
		pass
	#m = m+1
	#j = j+1
k=1	
for index, row in df.iterrows():
	#print(row['column_name'], row['data_type'], row['data_length'], row['nullable'])
	#if row['column_name']=='CUST_GRP_ID':
	#	for m in range(7,16):
	#		sheet.cell(row = m, column = j).value = '1'
	if row['data_type'] in ('VARCHAR', 'VARCHAR2', 'CHAR', 'NVARCHAR2'):
		a = int(row['data_length'])
		a1 = randomString(1)
		s1 = "'"+a1+"'"
		#print(s1)
		#sheet.write(3,j, "'I'")
		for m in range(16,25):
			if sheet.cell(row = 1, column = k).value == 'OP_CODE':
				sheet.cell(row = m, column = k).value = random.choice(loe)  #"'I'"
			else:
				sheet.cell(row = m, column = k).value = s1
		k = k+1
	elif row['data_type'] in ('NUMBER'):
		pr = 1 #int(sheet.cell(row = 3, column = j).value)
		sc = 1 #int(sheet.cell(row = 4, column = j).value)
		s2 = random_with_N_digits(pr)#'2'#'1' #random.randint(1, 1000) #random.sample(range(1, 10), 1)
		#print(s2)
		#sheet.write(3,j, s2)
		if row['column_name']=='CUST_GRP_ID':
			for m in range(16,25):
				sheet.cell(row = m, column = k).value = '2'
			k=k+1
		else:
			for m in range(16,25):
				#pr = int(sheet.cell(row = 3, column = j).value)
				#sc = int(sheet.cell(row = 4, column = j).value)
				s2 = random_with_N_digits(1)#'2'#'1' #random.randint(1, 1000) #random.sample(range(1, 10), 1)
				#print(s2)
				sheet.cell(row = m, column = k).value = s2 #m-3
			k = k+1
	elif row['data_type'] in ('DATE'):
		s3 = "'01-Jan-19'" #randomDate("1/1/2018", "1/2/2018", random.random())
		#print(s3)
		#sheet.write(3,j, s3)
		for m in range(16,25):
			sheet.cell(row = m, column = k).value = s3
		k = k+1
	elif 'TIMESTAMP' in row['data_type']:	#row['data_type'] like '%TIMESTAMP%':	#in ('TIMESTAMP'):
		#s4 = datime('2017-1-1 03:17:25 PM')
		s4 = "'"+"01-FEB-17 04.39.54.000000000 PM"+"'"
		#print(s4)
		#sheet.write(3,j, s4)
		for m in range(16,25):
			sheet.cell(row = m, column = k).value = s4
		k = k+1
	elif row['data_type'] in ('ROWID'):
		pass
	elif row['data_type'] in ('CLOB'):
		pass

	
con.close() 
#book.save("joinlist.xls")
wb.save("joinlist.xlsx")
	
wb = xlrd.open_workbook('joinlist.xlsx')
sheet = wb.sheet_by_index(0)

for p in range(6,24):
	half_query = 'INSERT INTO '+table+' ('
	with open("query.txt","w") as f:
		f.write(half_query)
		f.close()

	#for r in range(0, sheet.nrows):
	#	if sheet.cell(r, 0).value in ('091 - Fortnightly Dispute Sum Report','095 - Daily Report of Unset Blacklist','1000 - Adjustment to own account or family (to catch cheating staff)','1001 - Multiple Adjustment to Same Account (to catch cheating customer)','1002 - Deleted, Reversed, Corrected Charges Report','1003 - Deletion of Pending Cycle Events','1004 - Roaming Events Extraction','1005 - Late Charges From Ceased Line & Active Line Without Order','1006 - Extracted Rated Events Report','1007 - Daily eBill Sign-up and De-Registered Report','1008 - Monthly eBill Sign-up Report','1009 - Monthly eBill De-Registered Report','1010 - Monthly Consumer Report','1011 - Campaign Script Report','1012 - Repeated Interaction Report','1013 - Daily Payment Backout-Fund Transfer Activities Report','1014 - Monthly Deposit Ledger','1015 - Receipts Accumulated in ECA','1016 - FA Activities Ledger for the month','1017 - Monthly Statistic for Bills generated with BRE insert','1018 - Daily E-payment Details Report','1019 - Refund Summary by Dept','1020 - Daily Fraud Listing','1021 - Daily High Toll Listing','1022 - Monthly IS Listing','1023 - Monthly Minor and Foreigner Listing','1024 - Adhoc Mobile Application by Vendor and Dealer for Fraud Analysis','1025 - Fraud Analysis Report','1026 - Daily Interim Bill Successfully Generated Report','1027 - Monthly Involuntary Cessation','1028 - Age Analysis for all Cancelled Financial Accounts & Payment','1029 - Monthly Bad Debt and Cumulative WriteOffs Report','1030 - Quarterly Pending Write Off Report','1031 - Quarterly Authorised Write Off Report','1032 - Monthly Ageing Report for Consumer Residential','1033 - Monthly Mobile Fraud Statistic Report','1034 - Monthly Report of Cancelled Financial Account by Financial Year','1035 - Provision of Information (POI) for Telecommunication Bills in Arrears','1036 - Monthly Age Analysis for Current Subscribers (Credit)','1037 - Monthly Age Analysis for Historic Subscribers (Credit)','1038 - Monthly Age Analysis for Current Subscribers (Debit)','1039 - Monthly Age Analysis for Historic Subscribers (Debit)','1040 - Monthly Billings Report','1041 - Summary of Payment (Part 4)','1042 - AppleCare Plus Report','1043 - Daily Postpaid Porting','1044 - Postpaid Porting by Price Plan Mix','1045 - MobileShare Performance Detail Report','1046 - MobileShare Attachment Rate Report','1047 - Monthly SIM-Only Plan (Project Vanilla) Performance','1048 - WiVo Performance Report','1049 - WiVo Usage Consumption Report','1052 - DataX Performance Report','1053 - Quintet Attachment Rate Report','1054 - Daily Dash ROI','1055 a - Detailed Connection Report (Sales)','1055 b - Detailed Connection Report (Finance)','1056 - Smart Home Detail Report','1057 - Mobile Swop_SwopUp and Concierge Subbase report','1058 - Mobile SwopUp and Concierge Subscriber Report','1062 - List of All the Staff Rate Accounts Not on Giro','1063 - Conversion for Staff Rate not on Giro to Commercial Rate','1064 - Dash Pay Registration monthly Failure report','1065 - Dash Pay Registration Daily Report','1066 - Dash Pay Registration monthly Success report','1068 - Local registration requests','1069 - Foreigners registration requests','1071 - New activities related to Sim Only Plan','1076 - Monthly Project Circle Base report','1077 - Weekly Project Circle Base report','1078 - Network Filtering Service Performance','1079 - My Time Data Pack Performance Report','1080 - Connected cars','1081 - Singtel IPP Subscriber Profile','1082 - Singtel IPP Take-up Report','1083 - Singtel IPP Customers Installment Profile','1084 - Smart Home Removal Of Additional Sensor Service MRC Report','1085 - Singtel OTT Subbase report','1086 - Singtel OTT Performance report','1087 - Smart Home Daily Ops Report','1088 - Smart Home Subscription Base','1089 - Detailed Smart Home Report','1090 - Mobile SwopUp CI Eligible Base Report','1091 - Voucher Redemption Report','1251 - Mobile Threat Protection Subbase','131 - Postpaid Preactivation Subscribers','132 - Preactivated postpaid subscribers Order Actions','1595 - Monthly Deposit Release Review','1596 - Re-Issue of Direct Debit','1597 - Credit Rating and Foreigner Deposit','1598- Staff Rate Exception Report','1599 - Extraction of Component ID and description','1600 - Write off Accounts with Equipment Penalties Report','1601 - Monthly Debt Agent Performance Report','1602 - Audit Adjustment Report (Part 1 - Top CCOs Approvers Departments','1603 - Audit Adjustment Report (Part 2 - Exceeds Empowerment)','1604 - Monthly CallZone Report','1605 - Mobile MSF and AI Accruals','1607 - OP VAS','1609 - Audit Adjustment Report (Part 3 - Repeated Adj Type for Same Contact)','1632 - Billed Revenue By Country','1635 - Refund Report','1636 - Customer Deposit and Bill Payment Report','1675 - Summary of Payment (Part 5)','1676 - Monthly TD0 Report','1686 - Omni Channel Mobile Report','1687 - Spring D fraud Report – Address Customer‬','1688 - Eshop Delivery Report','1689 - OnePay for Telesales','1709 - Ambassador Consumer Sales Ops Report','1713 - Daily Overall MNP Report (BCC)','1714 - Weekly Cancelled Orders via Door to Door or RIM Sales Channels','1717 - Spring D fraud Report - Contact','1718 - Spring D fraud Report - High Level Plans','1720 - Orders in Negotiation Status raised via door to door or RIM sales channels','1722 - Near Real Time - Take Up Sales Report - Energy','1723 - Near Real Time - Stuck in Hold Activity Report - Energy','1730 - Near Real Time - Spring D promo code report','1735 - Near Real Time - Orders in Negotiation report','222 - Summary of Deposit by Subscription Type','239 - Credit Card Commission','246 - Daily Dishonoured Cheque','250 - Account By Payment Method & Subscription Type','251 - Monthly Payment Journal for HQ','260 - Monthly Analysis of Payments by Payment Modes and VIP Code','261 - Analysis of Successful and Unsuccessful GIRO - Cr Card Deductions','285 - Daily Audit Trail of Auto_Manual Refunds with Authorised Refund Details','286 - Unsuccessful Credit Card Deduction','288 - Direct Credit - Monthly Maintenance','290 - GIRO Maintenance','291 - Weekly Unsuccessful Deduction for EFT Accounts','296 - Summary Refund Report to Director','303 - Credit Card Maintenance','438 - Fortnightly Pause Report','439 - Daily Report of Termination Notices','441 - Monthly Report for STOP Collection','442 - Daily Accounts having PL-Restore-Cease Orders Pending','443 - Monthly Report For Waiver of Collection Fees','444 - Daily Report of Successful Involuntary Cancellations','446 - Daily Report of Frequent Manual Restorations','448 - Daily Report of Account Force-Stop Collection by Users','449 - Daily Report For Failed Terminations, Failed Cancellation & Pending Cancellation Orders','450 - Daily Report of Dishonoured Cheques','452 - Near Real Time - Detailed Sales Report - Singtel TV Content','452 - Near Real Time - Detailed Sales Report - Singtel TV Content_A','454 - Near Real Time - Detailed Sales Report - Mobile','454 - Near Real Time - Detailed Sales Report - Mobile_A','457 - Production Bill Check (By Billing Offers)','463 - Event Rejection Detailed Report','477 - Daily Report of Accounts Stuck in Collection Steps for 1 Month','503 - Customer change in bill cycle','507 - Monthly Billed Revenue for Detail Billing and Copy Bill Charges','508 - Success Held Back Bills (Hold Bill)','509 - Daily Reconciliation from URMS-Atomic-MMIG to OFCA-OCA to TC','510 - Near-Real-Time - Detailed Sales Report - Bundle','510 - Near-Real-Time - Detailed Sales Report - Bundle_A','511 - Near Real Time - Detailed Sales Report - VAS','511 - Near Real Time - Detailed Sales Report - VAS_A','512 - Bundle Query - mio Home Nomination Sales Query','513 - Near Real Time - Detailed Singtel TV Termination Report','514 - Movement Analysis of Unallocated Receipts in ECA','518 - Individual Detailed Refund Report - By Value Range','519 - Summary of Payment (Part 1)','520 - Summary of Payment by User & Manual File Upload','521 - Daily Summary of DD & CC Extractions','522 - Near-Real-Time - Backdated Order Report','523 - Near Real Time - Closed Cases Reason For Follow-up For SMS Poll','524 - Near Real Time Consumer Operations Dashboard','528 - Major Roadshow Sales Performance','529 - Daily-Monthly Movements Report for Payment Arrangements','530 - Monthly Analysis of Accounts with Payment Arrangements','538 - Daily Report of Audit Trail on Debit and Credit Balance Transfer','540 - Daily Report for Cancelled FA Payments','541 - Daily Payment for Written Off Accounts','542 - Near Real Time - Detailed Order Cancellation Report','543 - Near Real Time - Sales - Phone Launch Sales Performance','547 - Adjustment Report (Part 1 - Cross Departments)','552 - Provide Order Action SLA Report','553 - Monthly Report of Change Policy','554 a - Adjustments by RC Code (Approved and Pending)','554 b - Adjustments by RC Code (Approved and Pending)','554 c - Adjustments by RC Code (Approved and Pending)','555 - Audioline-Telepoll Credit NRC Adjustment Settlement','557 - Cessation of Services with Split Bill Arrangement','559 - Bill Cycle Summary','562 - Waiver Details Query','563 - Sales Conversion Report','564 - Monthly Working DEL Fixed line Status report - Traffic Usage','565 - Monthly Working DEL Fixed line Status report - Working Lines','566 - CallZone Tax Invoice Verification Against Pegasus','568 - Monthly Interconnect Settlement Report','569 - Daily Cross Carriage Reconciliation for requests sent out by SQL to RQL','570 - Daily Swap STB DIY Report','571 - Monthly Termination and Waiver for loss of STB','576 - Account with Changes in Bill Media','578 - Billing Rejection Report','579 - Rejected Service Numbers with Pending Orders','580 - Monthly Summary of Write-off Events','581 - Pending Events Rejections','582 - Rejected Events more than or equal to 90 days','583 - Usage Rejection for write-off and Ageing','584 - Analysis of Billed Accounts Market Codes and Services Type','585 - Daily OFCA to TC Event Reconciliation','588 - All pending Adjustments and aging','590 - Active Lines with no Meter Usage for 30 - 37 days','591 - Near Real Time - Detailed Sales Report - Cross Carriage','592 - Customer Rejection','593 - Monthly Termination Report for SPEAR dealer D2575','595 - Monthly Sales Transaction Count Report','597 - Summary of Payment (Part 2)','598 - Summary of Payment (Part 3)','599 - Near Real Time - Interaction Call Drivers Report','600 - Interaction Count Report','601 - Interaction Call Drivers Report','602 - Interaction Raw Data','603 - Near Real Time - Interaction Raw Data','606 - Adjustment Report (Part 2 - Amount and Reason Analysis)','610 - Order Action Type Count By Billing Offer Report','612 - Case Count Daily Report','614 - Order Action Count Weekly Report','615 - Order Action Count Monthly Report','616 - Order Action Cross Departments Report','617 - Open Order Action Snapshot Report','618 - Order Action Analysis Report','620 - Near Real Time - Case Type Report','621 - Open Case Snapshot Daily Report','622 - Open Case Within SLA Snapshot Daily Report','623 - Case Type Report','624 - Case Type By Product Report','625 - Closed Case Milestone Report','626 - Case Escalated To FDO Report','627 - Case Cross Departments Report','628 - Near Real Time - Callback Report','629 - Callback Report','630 - Near Real Time - Case Raw Data','631 - Case Raw Data','633 - Case Adjustment Raw Data','636 - Adjustment Report (Part 3 - Employee Analysis)','637 - Adjustment Report (Part 4 - Product Analysis)','638 - Case Dashboard','642 - Total Deposit','643 - Deposit Movement for the Month Report','644 - Daily Deposit Request-Cancel-Paid-Release Report','649 - Sales - ODS - Trade In Stock Report','650 - Staff Sales Report','652 - Activities Report for Consumer Residential_1','653 - Activities Report for Consumer Residential_2','653 - Daily Activities Report for Consumer Residential_2','654 - Activities for Consumer Residential Part_3','655 - Daily Projected Activities','656 - NRC Update via GUI & Mass Activities Batch upload Report','657 - Activities for Consumer Residential Details Reports','658 - Daily_Activities for Consumer Residential by Product Type','SMB - Autoactivation Delay in Order Provision','SMB - Autoactivation Delay in Order Provision_old'):
	with open("query.txt","a") as f:
		for i in range(0, sheet.ncols):	
			if sheet.cell(1, i).value!='TIMESTAMP' or sheet.cell(1, i).value!='CLOB':
				output = sheet.cell(0, i).value+','
				f.write(output)
			#f= open(sheet.cell(r, 1).value+".xml","w+")
			#f.write(sheet.cell(r, 4).value)
			#f.close()
	hq2 = ') values('
	with open("query.txt","a") as f:
		f.write(hq2)
		f.close()

	with open("query.txt","a") as f:
		for i in range(0, sheet.ncols):	
			if sheet.cell(1, i).value!='TIMESTAMP' or sheet.cell(1, i).value!='CLOB':
				output = str(sheet.cell(p, i).value)+','
				f.write(output)	

	hq3 = ');'
	with open("query.txt","a") as f:
		f.write(hq3)
		f.close()
		
	with open("query.txt","r") as f:
		qu = f.readlines()
		
	a='a'
	b='b'
	c='c'
	for line in qu:
		a,b,c = line.split(')')
		#print(str(line)+'\n\n')
		


	#a,b,c = qu.split(')')

	#print(a.rstrip(',')+'\n')
	#print(b.rstrip(',')+'\n')
	#print(c.rstrip(',')+'\n')

	d = a.rstrip(',')+')'+b.rstrip(',')+')'#+c.rstrip(',')
	#e = str(d).strip('[').d.strip(']')
	#print(d)

	with open("query1.txt","w+") as f:
		f.write(d)
		f.close()
		
	con = engine.connect()
	engine.execute(d)
	engine.execute('COMMIT')
	#df = pd.DataFrame(output.fetchall())
	#df.columns = output.keys()
	con.close() 
	


con.close() 


print("Done")